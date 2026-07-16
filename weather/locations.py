from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from grid import grid_to_latitude_longitude
from models import GridPoint, WeatherLocation


EXPECTED_HEADERS = (
    "구분",
    "행정구역코드",
    "1단계",
    "2단계",
    "3단계",
    "격자 X",
    "격자 Y",
    "경도(시)",
    "경도(분)",
    "경도(초)",
    "위도(시)",
    "위도(분)",
    "위도(초)",
    "경도(초/100)",
    "위도(초/100)",
    "위치업데이트",
)


@dataclass(frozen=True)
class LocationDataset:
    """위치 엑셀에서 읽은 고유 격자와 행정구역 목록이다."""

    grid_points: tuple[GridPoint, ...]
    locations: tuple[WeatherLocation, ...]


def load_location_dataset(path: Path) -> LocationDataset:
    """기상청 행정구역별 격자·위경도 XLSX를 검증해 읽는다."""

    if not path.is_file():
        raise FileNotFoundError(f"격자·위경도 파일을 찾을 수 없습니다: {path}")

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.worksheets[0]
        rows = worksheet.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration as exc:
            raise ValueError("격자·위경도 파일이 비어 있습니다.") from exc
        if tuple(header[: len(EXPECTED_HEADERS)]) != EXPECTED_HEADERS:
            raise ValueError("격자·위경도 파일의 헤더가 예상 형식과 다릅니다.")

        locations: list[WeatherLocation] = []
        grid_points: dict[tuple[int, int], GridPoint] = {}
        administrative_codes: set[str] = set()
        for row_number, row in enumerate(rows, start=2):
            if not row or all(value is None for value in row):
                continue
            location = _parse_location_row(row, row_number)
            if location.administrative_code in administrative_codes:
                raise ValueError(
                    f"행정구역코드가 중복되었습니다: {location.administrative_code}"
                )
            administrative_codes.add(location.administrative_code)
            locations.append(location)
            grid_points.setdefault(
                (location.nx, location.ny),
                grid_to_latitude_longitude(location.nx, location.ny),
            )
    finally:
        workbook.close()

    return LocationDataset(
        grid_points=tuple(grid_points[key] for key in sorted(grid_points)),
        locations=tuple(locations),
    )


def _parse_location_row(row: tuple[Any, ...], row_number: int) -> WeatherLocation:
    kind = _clean_text(row[0])
    if kind != "kor":
        raise ValueError(f"{row_number}행의 구분 값이 올바르지 않습니다: {kind!r}")

    administrative_code = _administrative_code(row[1], row_number)
    region_level_1 = _required_text(row[2], "1단계", row_number)
    region_level_2 = _optional_text(row[3])
    region_level_3 = _optional_text(row[4])
    if region_level_3 is not None and region_level_2 is None:
        raise ValueError(f"{row_number}행은 2단계 없이 3단계 지역명이 있습니다.")

    nx = _required_int(row[5], "격자 X", row_number)
    ny = _required_int(row[6], "격자 Y", row_number)
    grid_to_latitude_longitude(nx, ny)

    longitude = _required_float(row[13], "경도", row_number)
    latitude = _required_float(row[14], "위도", row_number)
    if longitude == 0.0 and latitude == 0.0:
        longitude = None
        latitude = None
    elif longitude == 0.0 or latitude == 0.0:
        raise ValueError(f"{row_number}행의 위경도 중 한 값만 0입니다.")
    elif not 120.0 <= longitude <= 140.0 or not 30.0 <= latitude <= 45.0:
        raise ValueError(f"{row_number}행의 위경도가 한국 영역을 벗어났습니다.")

    return WeatherLocation(
        administrative_code=administrative_code,
        region_level_1=region_level_1,
        region_level_2=region_level_2,
        region_level_3=region_level_3,
        nx=nx,
        ny=ny,
        longitude=longitude,
        latitude=latitude,
        source_updated_on=_optional_date(row[15], row_number),
    )


def _administrative_code(value: Any, row_number: int) -> str:
    if isinstance(value, int):
        normalized = f"{value:010d}"
    elif isinstance(value, float) and value.is_integer():
        normalized = f"{int(value):010d}"
    else:
        normalized = _clean_text(value)
    if len(normalized) != 10 or not normalized.isdigit():
        raise ValueError(f"{row_number}행의 행정구역코드 형식이 올바르지 않습니다.")
    return normalized


def _optional_date(value: Any, row_number: int) -> date | None:
    if value is None or _clean_text(value) == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _clean_text(value)
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    try:
        return datetime.strptime(text, "%Y%m%d").date()
    except ValueError as exc:
        raise ValueError(f"{row_number}행의 위치업데이트 날짜가 올바르지 않습니다.") from exc


def _required_text(value: Any, name: str, row_number: int) -> str:
    normalized = _clean_text(value)
    if not normalized:
        raise ValueError(f"{row_number}행의 {name} 값이 비어 있습니다.")
    return normalized


def _optional_text(value: Any) -> str | None:
    normalized = _clean_text(value)
    return normalized or None


def _required_int(value: Any, name: str, row_number: int) -> int:
    try:
        normalized = int(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"{row_number}행의 {name} 값이 정수가 아닙니다.") from exc
    return normalized


def _required_float(value: Any, name: str, row_number: int) -> float:
    try:
        return float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"{row_number}행의 {name} 값이 숫자가 아닙니다.") from exc


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()
